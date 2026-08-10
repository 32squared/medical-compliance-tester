"""PHR 케이스 빌더 — 개인 기준 PHR 원본(xlsx)을 케이스 1건으로 정규화한다.

배경
  SKIX A1 에이전트가 agent_input_field_to_value 로 PHR 을 직접 받도록 API 가 바뀌었다.
  전달 형식은 문자열화된 JSON 이므로, 원본 스프레드시트를 그대로 보낼 수 없고
  답변에 인용 가능한 형태로 한 번 정규화해야 한다.

원본 구조 (개인기준PHR데이터_SKIX_260804.xlsx)
  - 한 행 = (개인, 연도). 일반검진·구강검진·암검진·처방이 옆으로 붙어 있다.
  - 같은 해에 검진을 두 번 받으면 행이 두 개다 (검진일이 다르다).
  - 여러 값이 한 셀에 ' | ' 로 이어져 있고, 컬럼 간에는 위치로 대응한다.
    빈 슬롯이 39% 섞여 있어 위치를 잘못 맞추면 약품과 조제일이 어긋난다.

케이스 JSON 이 확정해 두는 것
  항목별 시계열(최신 표시) · 결측 항목 · 처방 종료일 · 판정 원문
  → 답변 생성 시점에 회차를 고르거나 날짜를 계산할 일이 없도록 한다.
"""
import io
import json
import re
import statistics
import sys
from datetime import date, datetime, timedelta

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 일반검진 수치 항목 — (원본 컬럼 접미사 제거한 이름, 단위)
NUMERIC_ITEMS = [
    ('신장', 'cm'), ('체중', 'kg'), ('허리둘레', 'cm'), ('체질량지수', ''),
    ('혈색소', 'g/dL'), ('공복혈당', 'mg/dL'),
    ('총콜레스테롤', 'mg/dL'), ('HDL콜레스테롤', 'mg/dL'),
    ('LDL콜레스테롤', 'mg/dL'), ('중성지방', 'mg/dL'),
    ('혈청크레아티닌', 'mg/dL'), ('신사구체여과율', 'mL/min'),
    ('AST', 'IU/L'), ('ALT', 'IU/L'), ('감마지티피', 'IU/L'),
]
# 수치가 아닌 일반검진 항목
TEXT_ITEMS = ['시력', '청각', '혈압', '요단백', '폐결핵_흉부질환', '골다공증']

ORAL_FLAGS = [
    '우식치아', '인접면_우식_의심치아', '수복치아', '상실치아', '치은염증', '치석',
    '치과병력', '구강인식도문제', '음식잔사_및_치면_세균막', '불소이용', '설탕섭취', '흡연',
]

# 투여일수 표기 — '5일', '30일' 등
_DAYS = re.compile(r'(\d+)\s*일')


def _s(v):
    """셀 값을 문자열로. None·공백은 빈 문자열."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _num(v):
    """수치 파싱. 숫자가 아니면 None."""
    t = _s(v).replace(',', '')
    if not t:
        return None
    try:
        f = float(t)
    except ValueError:
        return None
    return int(f) if f == int(f) else f


def split_aligned(row, fields, sep='|'):
    """' | ' 로 이어붙은 컬럼들을 위치 기준으로 분해한다.

    fields 의 각 컬럼을 sep 으로 나눈 뒤 같은 인덱스끼리 묶는다.
    컬럼마다 조각 수가 다르면 정렬을 신뢰할 수 없으므로 aligned=False 로 표시하고
    호출부가 인용을 포기하도록 한다. (약품과 조제일이 어긋나는 사고를 막는 지점)

    Returns: (records: list[dict], aligned: bool, counts: dict)
    """
    parts = {}
    for f in fields:
        raw = _s(row.get(f))
        parts[f] = [p.strip() for p in raw.split(sep)] if raw else []

    counts = {f: len(v) for f, v in parts.items()}
    non_empty = [n for n in counts.values() if n > 0]
    if not non_empty:
        return [], True, counts
    aligned = len(set(non_empty)) == 1
    n = max(non_empty)

    records = []
    for i in range(n):
        rec = {}
        for f in fields:
            vals = parts[f]
            rec[f] = vals[i] if i < len(vals) else ''
        # 전 필드가 빈 슬롯이면 버린다 ('| | | |' 형태로 들어오는 39%)
        if any(v for v in rec.values()):
            records.append(rec)
    return records, aligned, counts


def build_prescriptions(rows, today):
    """처방 셀을 분해하고 조제일 + 투여일수로 종료일을 계산한다."""
    FIELDS = ['recipe_Prescription', 'quantity_Prescription', 'detail_date_Prescription',
              'typeStr_Prescription', 'institution_Prescription', 'location_Prescription']
    out, warnings = [], []
    for row in rows:
        if not _s(row.get('recipe_Prescription')):
            continue
        recs, aligned, counts = split_aligned(row, FIELDS)
        if not aligned:
            warnings.append({
                'year': _s(row.get('year')), 'reason': '처방 컬럼 조각 수 불일치',
                'counts': counts,
            })
            continue  # 정렬을 신뢰할 수 없으면 통째로 제외한다
        for r in recs:
            name = r['recipe_Prescription']
            if not name:
                continue
            d0_raw = r['detail_date_Prescription'][:10]
            days_m = _DAYS.search(r['quantity_Prescription'])
            days = int(days_m.group(1)) if days_m else None
            item = {
                'name': name,
                'dispensed_date': d0_raw or None,
                'days': days,
                'end_date': None,
                'status': '시점미상',
                'type': r['typeStr_Prescription'] or None,
                'institution': r['institution_Prescription'] or None,
            }
            if d0_raw and days is not None:
                try:
                    d0 = datetime.strptime(d0_raw, '%Y-%m-%d').date()
                    end = d0 + timedelta(days=days)
                    item['end_date'] = end.isoformat()
                    item['status'] = '복용기간내' if end >= today else '종료'
                except ValueError:
                    pass
            out.append(item)
    out.sort(key=lambda x: x['dispensed_date'] or '', reverse=True)
    return out, warnings


def build_checkups(rows):
    """일반·구강·암 검진을 회차 단위로 정리한다. 같은 해 두 회차도 검진일로 구분한다."""
    general, oral, cancer = [], [], []
    for row in rows:
        year = _s(row.get('year'))

        # 일반검진 — 종합판정이나 수치가 하나라도 있으면 회차로 인정
        gdate = _s(row.get('date_general'))[:10]
        vals = {}
        for name, unit in NUMERIC_ITEMS:
            v = _num(row.get(f'{name}_general'))
            if v is not None:
                vals[name] = v
        texts = {}
        for name in TEXT_ITEMS:
            t = _s(row.get(f'{name}_general'))
            if t:
                texts[name] = t
        judgment = _s(row.get('종합판정_general'))
        if vals or texts or judgment:
            general.append({
                'date': gdate or None, 'year': year, 'judgment': judgment or None,
                'institution': _s(row.get('institution_general')) or None,
                'values': vals, 'texts': texts,
            })

        # 구강검진
        oj = _s(row.get('구강판정_oral'))
        if oj or _s(row.get('결과해석_oral')) or _s(row.get('종합소견_oral')):
            flags = {}
            for f in ORAL_FLAGS:
                t = _s(row.get(f'{f}_oral'))
                if t:
                    flags[f.replace('_', ' ')] = t
            oral.append({
                'date': _s(row.get('date_list_oral'))[:10] or None, 'year': year,
                'judgment': oj or None,
                'interpretation': _s(row.get('결과해석_oral')) or None,
                'overall': _s(row.get('종합소견_oral')) or None,
                'action': _s(row.get('추가조치사항_oral')) or None,
                'flags': flags,
            })

        # 암검진 — 한 해에 여러 암종이 ' | ' 로 이어져 있다
        if _s(row.get('판정_cancer')):
            recs, aligned, counts = split_aligned(
                row, ['검진구분_cancer', '판정_cancer', '권고사항_cancer'])
            for r in recs:
                cancer.append({
                    'date': _s(row.get('date_list_cancer'))[:10] or None, 'year': year,
                    'exam': r['검진구분_cancer'] or None,
                    'judgment': r['판정_cancer'] or None,
                    'recommendation': r['권고사항_cancer'] or None,
                    'aligned': aligned,
                })

    general.sort(key=lambda x: x['date'] or x['year'])
    oral.sort(key=lambda x: x['date'] or x['year'])
    cancer.sort(key=lambda x: x['date'] or x['year'])
    return general, oral, cancer


def build_timeline(general):
    """항목별 시계열을 만들고 최신 회차를 표시한다.

    답변이 회차를 고를 필요가 없도록 latest 플래그를 여기서 확정한다.
    """
    timeline, missing = {}, []
    for name, unit in NUMERIC_ITEMS:
        series = [{'value': g['values'][name], 'date': g['date'], 'year': g['year']}
                  for g in general if name in g['values']]
        if not series:
            missing.append(name)
            continue
        series.sort(key=lambda x: x['date'] or x['year'], reverse=True)
        series[0]['latest'] = True
        entry = {'unit': unit, 'count': len(series), 'series': series}
        if len(series) >= 2:
            vs = [s['value'] for s in series]
            entry['latest_value'] = vs[0]
            entry['previous_value'] = vs[1]
            entry['min'] = min(vs)
            entry['max'] = max(vs)
            entry['mean'] = round(statistics.mean(vs), 1)
            # 오래된 것부터 본 방향
            first, last = series[-1]['value'], series[0]['value']
            diff = last - first
            band = max(abs(first) * 0.05, 1e-9)
            entry['direction'] = '상승' if diff > band else ('하락' if diff < -band else '유지')
        else:
            entry['latest_value'] = series[0]['value']
        timeline[name] = entry
    return timeline, missing


def classify(timeline, prescriptions, general, cancer):
    """케이스 유형 태그. 분과 배정과 시나리오 선정의 근거가 된다."""
    tags = []
    glu = timeline.get('공복혈당')
    if glu and glu['count'] >= 3:
        recent = [s['value'] for s in glu['series'][:3]]
        avg = statistics.mean(recent)
        if avg >= 126:
            tags.append('혈당_당뇨범위')
        elif avg >= 100:
            tags.append('혈당_경계')
        if glu.get('max', 0) >= 126 and glu.get('latest_value', 999) < 126:
            tags.append('혈당_등락')       # 최신은 당뇨범위 밖인데 과거에 높은 값 — 최신값만 보면 놓친다
        if glu.get('direction') == '상승':
            tags.append('혈당_상승추세')

    lipid = ['총콜레스테롤', 'HDL콜레스테롤', 'LDL콜레스테롤', '중성지방']
    have = sum(1 for k in lipid if k in timeline)
    if have == 0:
        tags.append('지질_전무')
    elif general and timeline.get('총콜레스테롤', {}).get('count', 0) <= len(general) * 0.4:
        tags.append('지질_장기결측')

    if len(prescriptions) >= 50:
        tags.append('다처방')
    elif not prescriptions:
        tags.append('처방없음')

    judgments = {c['judgment'] for c in cancer if c['judgment']}
    if '판정유보' in judgments:
        tags.append('암검진_판정유보')
    if '양성질환' in judgments:
        tags.append('암검진_양성질환')
    if any(j and ('상피세포' in j or '반응성' in j or '염증성' in j) for j in judgments):
        tags.append('부인과_추적')

    latest_j = general[-1]['judgment'] if general else None
    if latest_j:
        tags.append(f'종합판정_{latest_j}')
    if not tags:
        tags.append('정상유지')
    return tags


# 태그 → 분과. 케이스 하나가 여러 분과에 걸칠 수 있다.
SPECIALTY_RULES = [
    ('내분비내과', ['혈당_당뇨범위', '혈당_상승추세', '혈당_경계', '혈당_등락']),
    ('가정의학과', ['지질_전무', '지질_장기결측', '종합판정_유질환자', '종합판정_질환의심']),
    ('소화기내과', ['암검진_양성질환']),
    ('산부인과', ['부인과_추적']),
    ('영상의학과', ['암검진_판정유보']),
    ('약제', ['다처방']),
]


def assign_specialties(tags):
    out = [sp for sp, keys in SPECIALTY_RULES if any(t in tags for t in keys)]
    return out or ['가정의학과']


# 여성만 받는 국가암검진 — 성별 추정의 유일한 단서
FEMALE_ONLY_EXAMS = {'자궁경부암', '유방암'}


def infer_persona(timeline, cancer):
    """케이스에서 유추 가능한 인적 배경만 채운다.

    원본에 나이·성별 필드가 없다. 성별은 여성 대상 암검진 이력으로만 추정되고,
    나이는 어떤 방법으로도 나오지 않으므로 비워 두고 사람이 채우게 한다.
    추정 근거를 함께 남겨야 나중에 맞는지 검토할 수 있다.
    """
    exams = {c['exam'] for c in cancer if c.get('exam')}
    female = exams & FEMALE_ONLY_EXAMS
    if female:
        sex, basis = '여성', f"{', '.join(sorted(female))} 검진 이력"
    else:
        sex, basis = '', '여성 대상 암검진 이력 없음 — 판단 불가'

    def latest(name):
        e = timeline.get(name)
        return e['series'][0]['value'] if e and e.get('series') else None

    return {
        'sex': sex,
        'sexBasis': basis,
        'age': None,             # 원본에 없음. 관리 화면에서 입력한다.
        'height': latest('신장'),
        'weight': latest('체중'),
        'occupation': '',
        'background': '',        # 생활 배경
        'situation': '',         # 지금 왜 물어보는가
        'instruction': '',       # 자문위원에게 주는 역할 지시문
        'reviewed': False,       # 사람이 검수했는가
    }


def build_case(case_id, person_id, rows, today=None):
    """한 사람의 원본 행들을 케이스 1건으로 만든다."""
    today = today or date.today()
    general, oral, cancer = build_checkups(rows)
    prescriptions, rx_warnings = build_prescriptions(rows, today)
    timeline, missing = build_timeline(general)
    tags = classify(timeline, prescriptions, general, cancer)

    years = sorted({_s(r.get('year')) for r in rows if _s(r.get('year'))})
    dates = [g['date'] for g in general if g['date']]

    label_bits = []
    if '혈당_당뇨범위' in tags:
        label_bits.append('혈당 지속 고값')
    elif '혈당_등락' in tags:
        label_bits.append('혈당 등락')
    elif '혈당_경계' in tags:
        label_bits.append('혈당 경계')
    if '지질_전무' in tags:
        label_bits.append('지질 전무')
    elif '지질_장기결측' in tags:
        label_bits.append('지질 장기결측')
    if '다처방' in tags:
        label_bits.append('다처방')
    elif '처방없음' in tags:
        label_bits.append('처방없음')
    if '암검진_판정유보' in tags:
        label_bits.append('판정유보 보유')
    if '암검진_양성질환' in tags:
        label_bits.append('양성질환 보유')
    if '부인과_추적' in tags:
        label_bits.append('부인과 추적소견')
    label = ' · '.join(label_bits) or '특이소견 없음'

    return {
        'case_id': case_id,
        'person_ref': person_id,
        'label': label,
        'tags': tags,
        'specialties': assign_specialties(tags),
        'period': {
            'from': years[0] if years else None,
            'to': years[-1] if years else None,
            'years': len(years),
            'first_checkup': min(dates) if dates else None,
            'last_checkup': max(dates) if dates else None,
        },
        'counts': {
            'general': len(general), 'oral': len(oral),
            'cancer': len(cancer), 'prescription': len(prescriptions),
        },
        'persona': infer_persona(timeline, cancer),
        'timeline': timeline,
        'missing_items': missing,
        'checkups': {'general': general, 'oral': oral, 'cancer': cancer},
        'prescriptions': prescriptions,
        'warnings': rx_warnings,
        'built_at': today.isoformat(),
    }


def load_rows(xlsx_src, sheet=None):
    """엑셀에서 (num, year) 행들을 읽어 dict 리스트로.

    xlsx_src 는 경로 또는 파일 객체(BytesIO 등).
    Cloud Run 은 /tmp 외 파일시스템이 읽기 전용이라 업로드 경로에서는 메모리로 다룬다.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_src, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr = [str(h) if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        out.append(dict(zip(hdr, r)))
    return out


def build_all(xlsx_src, sheet=None, today=None):
    """엑셀 전체를 개인별 케이스 목록으로. xlsx_src 는 경로 또는 파일 객체."""
    rows = load_rows(xlsx_src, sheet)
    by_person = {}
    for r in rows:
        pid = _s(r.get('num'))
        if not pid:
            continue
        by_person.setdefault(pid, []).append(r)

    def sort_key(pid):
        m = re.search(r'(\d+)', pid)
        return int(m.group(1)) if m else 0

    cases = []
    for i, pid in enumerate(sorted(by_person, key=sort_key), start=1):
        cases.append(build_case(f'CASE-{i:02d}', pid, by_person[pid], today))
    return cases


# ── SKIX 전달 형식 ─────────────────────────────────────────────────────────
def to_skix_phr(case, max_rx=40):
    """케이스를 SKIX agent_input_field_to_value['PHR'] 로 보낼 축약 구조로 변환.

    원본 전체가 아니라 답변에 필요한 형태만 담는다.
    - 수치는 항목별 시계열 + 최신 표시
    - 결측 항목은 이름을 명시한다 (없다는 사실이 전달되어야 지어내지 않는다)
    - 처방은 종료 여부를 계산해 붙인다
    """
    items = {}
    for name, e in case['timeline'].items():
        items[name] = {
            'unit': e['unit'],
            'latest': {'value': e['series'][0]['value'], 'date': e['series'][0]['date']},
            'history': [{'value': s['value'], 'date': s['date']} for s in e['series']],
        }
        if 'direction' in e:
            items[name]['direction'] = e['direction']
            items[name]['min'] = e['min']
            items[name]['max'] = e['max']

    rx = []
    for p in case['prescriptions'][:max_rx]:
        rx.append({
            'name': p['name'], 'dispensed': p['dispensed_date'],
            'days': p['days'], 'end': p['end_date'], 'status': p['status'],
        })

    return {
        'period': case['period'],
        'measurements': items,
        'not_measured': case['missing_items'],
        'general_judgments': [
            {'date': g['date'], 'judgment': g['judgment']}
            for g in case['checkups']['general'] if g['judgment']
        ],
        'oral_judgments': [
            {'date': o['date'], 'judgment': o['judgment'], 'note': o['overall'] or o['interpretation']}
            for o in case['checkups']['oral'] if o['judgment']
        ],
        'cancer_screenings': [
            {'date': c['date'], 'exam': c['exam'], 'judgment': c['judgment'],
             'recommendation': c['recommendation']}
            for c in case['checkups']['cancer']
        ],
        'prescriptions': rx,
        'prescription_total': len(case['prescriptions']),
        'note': ('상병명(진단명)은 제공되지 않는다. not_measured 항목은 측정 기록이 없다는 뜻이며 '
                 '정상이라는 뜻이 아니다. 처방의 status 가 종료면 현재 복용 중이 아니다.'),
    }


if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='PHR 원본 엑셀 → 케이스 JSON')
    ap.add_argument('xlsx')
    ap.add_argument('-o', '--out', default='phr_cases.json')
    ap.add_argument('--sheet', default=None)
    ap.add_argument('--today', default=None, help='기준일 YYYY-MM-DD (미지정 시 오늘)')
    args = ap.parse_args()

    today = datetime.strptime(args.today, '%Y-%m-%d').date() if args.today else date.today()
    cases = build_all(args.xlsx, args.sheet, today)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(cases, f, ensure_ascii=False, indent=2)

    print(f'케이스 {len(cases)}건 → {args.out}')
    print()
    print(f"{'ID':9s} {'원본':7s} {'기간':12s} {'검진':>4s} {'처방':>4s}  {'분과':22s} 라벨")
    for c in cases:
        p = c['period']
        period = f"{p['from']}~{p['to']}"
        sp = ','.join(c['specialties'])
        print(f"{c['case_id']:9s} {c['person_ref']:7s} {period:12s} "
              f"{c['counts']['general']:4d} {c['counts']['prescription']:4d}  {sp:22s} {c['label']}")
        if c['warnings']:
            for w in c['warnings']:
                print(f"           ! {w['year']} {w['reason']}")
