# -*- coding: utf-8 -*-
"""SKIX 2차 의료자문 준비 데이터 투입.

렉스소프트가 전달한 7개 분과 70명 PHR 엑셀과, 그 데이터에서 도출한 고정 문항을
시스템에 넣어 자문 당일 바로 진행할 수 있는 상태로 만든다.

투입 대상 네 가지
  1. PHR 케이스 70건    — 분과별 시트를 사람 단위 케이스로 빌드해 저장
  2. 자문위원 계정 7개  — rexsoft01~07 (이미 있으면 건드리지 않음)
  3. 케이스 배정        — 분과별 10명씩, 그 분과 자문위원에게만 보이도록
  4. 고정 문항 48건     — 6분과 × 6문항 + 응급의학 4문항 × 가상 Vital 3세트

문항은 reports/SKIX_2차자문_체크케이스_v*.xlsx 의 `01_체크문항` 시트를 읽는다.
그 파일이 문항의 단일 진실 소스이고 원본 585행과 대조 검증까지 끝난 산출물이다.

실행:
    python scripts/seed_advisory.py --phr <70명.xlsx> --questions <체크케이스.xlsx>
    python scripts/seed_advisory.py ... --dry-run     # 저장 없이 계획만 출력

같은 케이스 번호는 덮어쓰고, 같은 id 의 문항도 덮어쓴다. 여러 번 돌려도 결과는 같다.
"""
import argparse
import hashlib
import io
import os
import re
import secrets
import sys
import unicodedata

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import db  # noqa: E402
import phr_case_builder as pcb  # noqa: E402

# 시트 이름 → 자문 분과. 선정요약은 케이스가 아니라 배정 근거표라 제외한다.
SHEET_SPECIALTY = {
    '01. 가정의학과': '가정의학',
    '02. 내분비내과': '내분비내과',
    '03. 응급의학과': '응급의학',
    '04. 소화기내과': '소화기내과',
    '05. 순환기내과': '순환기내과',
    '06. 산부인과': '산부인과',
    '07. 유방암센터(영상의학과)': '영상의학',
}
# 문항 시트의 분과 표기가 시트명과 다르다 (엑셀은 짧은 이름을 쓴다)
QUESTION_SPECIALTY = {
    '가정의학': '가정의학', '내분비': '내분비내과', '응급의학': '응급의학',
    '소화기': '소화기내과', '순환기': '순환기내과', '산부인과': '산부인과',
    '영상의학': '영상의학',
}
# 분과 → 자문위원 계정. 렉스소프트 회신으로 배정이 바뀌면 이 표만 고치면 된다.
SPECIALTY_ADVISOR = {
    '가정의학': 'rexsoft01', '내분비내과': 'rexsoft02', '응급의학': 'rexsoft03',
    '소화기내과': 'rexsoft04', '순환기내과': 'rexsoft05', '산부인과': 'rexsoft06',
    '영상의학': 'rexsoft07',
}
SPECIALTY_SLUG = {
    '가정의학': 'FM', '내분비내과': 'EN', '응급의학': 'EM', '소화기내과': 'GI',
    '순환기내과': 'CV', '산부인과': 'OB', '영상의학': 'BR',
}
RISK_MAP = {'최고': 'CRITICAL', '높음': 'HIGH', '중간': 'MEDIUM', '낮음': 'LOW'}
# 응급 문항은 같은 질문을 세 세트에 던져 어느 구간부터 119 안내가 나오는지 본다
VITAL_SETS = ['V-A', 'V-B', 'V-C']


def _hash_password(password, salt=None):
    """proxy_server._hash_password 와 동일 규칙 (pbkdf2_hmac sha256, 100k)."""
    if salt is None:
        salt = secrets.token_hex(16)
    pw_hash = hashlib.pbkdf2_hmac(
        'sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000).hex()
    return pw_hash, salt


def _nfc(s):
    return unicodedata.normalize('NFC', str(s or ''))


def find_file(pattern, roots):
    """한글 파일명은 NFC/NFD 차이로 glob 이 놓치므로 walk 하며 정규화 비교한다."""
    pat = re.compile(pattern)
    for root in roots:
        if not os.path.isdir(root):
            continue
        for dirpath, _dirnames, filenames in os.walk(root):
            for fn in filenames:
                if pat.search(_nfc(fn)):
                    return os.path.join(dirpath, fn)
    return None


# ════════════════════════════════════════════════════════════════════════════
# 1. PHR 케이스
# ════════════════════════════════════════════════════════════════════════════
def build_cases(xlsx_path, today=None):
    """분과별 시트를 훑어 케이스 목록을 만든다.

    case_no 는 분과 순서대로 전역 통번(CASE-01~CASE-70)을 다시 매긴다.
    build_all 은 시트마다 CASE-01 부터 시작하므로 그대로 두면 시트끼리 번호가 겹쳐
    나중 시트가 앞 시트를 덮어쓴다.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheets = {_nfc(n): n for n in wb.sheetnames}
    wb.close()

    cases, seen_person, n = [], {}, 0
    for want, specialty in SHEET_SPECIALTY.items():
        real = sheets.get(_nfc(want))
        if not real:
            print(f'  [건너뜀] 시트 없음: {want}')
            continue
        built = pcb.build_all(xlsx_path, sheet=real, today=today)
        for c in built:
            pid = c.get('person_ref', '')
            if pid in seen_person:
                # 같은 사람이 두 분과에 걸치면 케이스는 하나만 두고 분과만 덧붙인다
                prev = seen_person[pid]
                if specialty not in prev['specialties']:
                    prev['specialties'].append(specialty)
                print(f'  [중복] {pid} — {prev["case_id"]} 에 분과 {specialty} 추가')
                continue
            n += 1
            c['case_id'] = f'CASE-{n:02d}'
            # 분과는 태그 추정이 아니라 렉스소프트 배정을 따른다
            c['specialties'] = [specialty]
            c['payload'] = pcb.to_skix_phr(c)
            seen_person[pid] = c
            cases.append(c)
        print(f'  {specialty:<8} {len(built)}명')
    return cases


# ════════════════════════════════════════════════════════════════════════════
# 2. 문항
# ════════════════════════════════════════════════════════════════════════════
def read_questions(xlsx_path):
    """체크케이스 엑셀의 01_체크문항 시트를 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if '01_체크문항' not in wb.sheetnames:
        raise SystemExit(f'01_체크문항 시트가 없습니다: {xlsx_path}')
    ws = wb['01_체크문항']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(h or '').strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        out.append({h: (v if v is not None else '') for h, v in zip(hdr, r)})
    return out


def build_scenarios(questions, person_to_case):
    """문항 → 시나리오. 응급의학은 Vital 세트마다 1건씩 만든다."""
    made, skipped = [], []
    for q in questions:
        spec_raw = str(q.get('분과', '')).strip()
        specialty = QUESTION_SPECIALTY.get(spec_raw, spec_raw)
        code = str(q.get('문항', '')).strip()
        person = str(q.get('케이스 ID', '')).strip()
        prompt = str(q.get('질문 문장', '')).strip()
        if not prompt:
            continue
        case_id = person_to_case.get(person)
        if not case_id:
            skipped.append(f'{specialty} {code} — 케이스 미등록 {person}')
            continue

        item = str(q.get('대응 검토항목', '')).strip()
        tags = ['자문2차', f'분과:{specialty}', f'문항:{code}']
        if item:
            tags.append(f'검토항목:{item[:1]}')
        base = {
            'category': 'emergency' if specialty == '응급의학' else 'general',
            'subcategory': specialty,
            'prompt': prompt,
            'expectedBehavior': str(q.get('검토 포인트', '')).strip(),
            'riskLevel': RISK_MAP.get(str(q.get('위험도', '')).strip(), 'MEDIUM'),
            'shouldRefuse': False,
            'source': 'advisory',
            'enabled': True,
            'phrCaseId': case_id,
        }
        slug = SPECIALTY_SLUG.get(specialty, 'XX')
        # 연도 칸이 Vital 지정을 겸한다.
        #   'V-A/B/C' → 같은 질문을 세 세트에 던져 응급 개시선을 관측 (문항 1개 → 실행 3건)
        #   'V-B'     → 그 한 세트만 (순환기 Q3 처럼 특정 구간을 겨냥한 문항)
        year_cell = str(q.get('연도', '')).strip()
        if year_cell == 'V-A/B/C':
            vitals = list(VITAL_SETS)
        elif year_cell in VITAL_SETS:
            vitals = [year_cell]
        else:
            vitals = ['']
        for v in vitals:
            s = dict(base)
            s['phrVitals'] = v
            s['id'] = f'ADV-{slug}-{code}' + (f'-{v.replace("-", "")}' if v else '')
            s['tags'] = tags + ([f'Vital:{v}'] if v else [])
            made.append(s)
    return made, skipped


# ════════════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--phr', help='70명 PHR 엑셀 경로')
    ap.add_argument('--questions', help='체크케이스 엑셀 경로 (01_체크문항 시트)')
    # 비밀번호는 저장소에 남기지 않는다 — 계정을 새로 만들 때만 필요하고,
    # 이미 있는 계정은 건드리지 않으므로 재실행 시에는 없어도 된다.
    ap.add_argument('--password', default=os.environ.get('ADVISOR_PASSWORD', ''),
                    help='자문위원 계정 비밀번호 (신규 생성 시에만 필요, 환경변수 ADVISOR_PASSWORD 도 가능)')
    ap.add_argument('--dry-run', action='store_true', help='저장하지 않고 계획만 출력')
    args = ap.parse_args()

    roots = [os.getcwd(), os.path.expanduser('~/Desktop'), os.path.expanduser('~/Downloads'),
             os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reports'),
             r'C:\Users\20002652\project\medical-compliance-tester\reports']
    phr_path = args.phr or find_file(r'분과별_PHR.*\.xlsx$', roots)
    q_path = args.questions or find_file(r'체크케이스_v\d+\.xlsx$', roots) \
        or find_file(r'체크케이스\.xlsx$', roots)
    if not phr_path or not os.path.isfile(phr_path):
        raise SystemExit('PHR 엑셀을 찾지 못했습니다. --phr 로 지정하세요.')
    if not q_path or not os.path.isfile(q_path):
        raise SystemExit('체크케이스 엑셀을 찾지 못했습니다. --questions 로 지정하세요.')

    print('=' * 78)
    print('SKIX 2차 자문 데이터 투입' + (' [DRY RUN]' if args.dry_run else ''))
    print('=' * 78)
    print(f'PHR      {phr_path}')
    print(f'문항     {q_path}\n')

    db.init_db()

    # ── 0. 투입 전 현재 상태 ──
    # 같은 case_no 는 덮어쓰므로, 운영 DB 처럼 이미 데이터가 있는 곳에서는
    # 무엇을 밀어내는지 먼저 보여 준다.
    print('■ 0. 투입 전 현재 상태')
    try:
        before = db.get_phr_cases()
        print(f'  기존 PHR 케이스 {len(before)}건'
              + (f" ({before[0].get('caseNo')}~{before[-1].get('caseNo')})" if before else ''))
        srcs = {}
        for c in before:
            srcs[c.get('sourceFile') or c.get('source_file') or '-'] = \
                srcs.get(c.get('sourceFile') or c.get('source_file') or '-', 0) + 1
        for k, v in srcs.items():
            print(f'    출처 {k}: {v}건')
        adv = [s for s in db.get_scenarios_summary(light=True).get('scenarios', [])
               if s.get('source') == 'advisory']
        print(f'  기존 자문 문항 {len(adv)}건 (같은 id 는 갱신됨)')
    except Exception as e:
        print(f'  [확인 실패] {str(e)[:120]}')
    print()

    # ── 1. 케이스 ──
    print('■ 1. PHR 케이스')
    cases = build_cases(phr_path)
    print(f'  → 케이스 {len(cases)}건')
    person_to_case = {c['person_ref']: f"phr_{c['case_id']}" for c in cases}
    no_persona = [c['case_id'] for c in cases
                  if not (c.get('persona') or {}).get('age')]
    if not args.dry_run:
        db.save_phr_cases(cases, source_file=os.path.basename(phr_path), created_by='seed_advisory')
        print(f'  저장 완료')

    # ── 2. 계정 ──
    print('\n■ 2. 자문위원 계정')
    created, existing, skipped_pw = [], [], []
    for spec, uid in SPECIALTY_ADVISOR.items():
        try:
            if db.get_user(uid):
                existing.append(uid)
                continue
        except Exception:
            pass
        # 빈 비밀번호로 계정을 만들면 아무나 로그인할 수 있다. 만들지 않고 넘긴다.
        if not args.password:
            skipped_pw.append(uid)
            continue
        if not args.dry_run:
            pw_hash, salt = _hash_password(args.password)
            db.create_user({
                'id': uid, 'name': f'{spec} 자문위원', 'org': 'REX Soft', 'uid': '',
                'password_hash': pw_hash, 'password_salt': salt,
                'status': 'approved', 'role': 'advisor',
            })
        created.append(uid)
    print(f'  신규 {len(created)}건 {created}')
    print(f'  기존 {len(existing)}건 {existing}')
    if skipped_pw:
        print(f'  [건너뜀] 비밀번호 미지정 {len(skipped_pw)}건 {skipped_pw}'
              f' — --password 또는 환경변수 ADVISOR_PASSWORD 로 준다')

    # ── 3. 문항 (배정보다 먼저 — 문항이 배정 대상을 결정한다) ──
    print('\n■ 3. 고정 문항')
    questions = read_questions(q_path)
    scenarios, skipped = build_scenarios(questions, person_to_case)
    print(f'  원본 {len(questions)}행 → 시나리오 {len(scenarios)}건')
    for s in skipped:
        print(f'  [건너뜀] {s}')

    # ── 4. 배정 ──
    #
    # 분과 소속만으로 배정하면 안 된다. 어떤 분과의 문항이 다른 분과 시트의 사람을
    # 가리키는 경우가 있는데(예: 영상의학 문항이 소화기 시트의 ID_421 을 쓴다),
    # 그 케이스가 문항 담당 자문위원에게 배정되지 않으면 본인 문항이 화면에 뜨지 않는다.
    # 그래서 '분과 소속' 과 '문항이 참조함' 두 근거를 합쳐 배정한다.
    print('\n■ 4. 케이스 배정')
    assign = {}      # case_id → {uid: 배정 사유}
    for c in cases:
        cid = f"phr_{c['case_id']}"
        for spec in c['specialties']:
            uid = SPECIALTY_ADVISOR.get(spec)
            if uid:
                assign.setdefault(cid, {})[uid] = spec
    cross = []
    for s in scenarios:
        uid = SPECIALTY_ADVISOR.get(s['subcategory'])
        cid = s.get('phrCaseId')
        if uid and cid and uid not in assign.get(cid, {}):
            assign.setdefault(cid, {})[uid] = s['subcategory']
            cross.append(f"{cid.replace('phr_', '')} → {uid}({s['subcategory']})")
    if not args.dry_run:
        for cid, who in assign.items():
            db.set_phr_assignments(cid, list(who.keys()),
                                   specialty=next(iter(who.values())),
                                   assigned_by='seed_advisory')
    per_uid = {}
    for cid, who in assign.items():
        for uid in who:
            per_uid[uid] = per_uid.get(uid, 0) + 1
    for spec, uid in SPECIALTY_ADVISOR.items():
        print(f'  {spec:<8} {per_uid.get(uid, 0):>2}건 → {uid}')
    if cross:
        print(f'  분과 밖 문항 때문에 추가 배정 {len(cross)}건')
        for x in sorted(set(cross)):
            print(f'    {x}')

    print('\n■ 5. 문항 저장')
    if not args.dry_run:
        ok = fail = 0
        for s in scenarios:
            try:
                if db.get_scenario(s['id']):
                    db.update_scenario(s['id'], s)
                else:
                    db.create_scenario(s)
                ok += 1
            except Exception as e:
                fail += 1
                print(f"  [실패] {s['id']}: {str(e)[:120]}")
        print(f'  저장 {ok}건 / 실패 {fail}건')
    by_spec = {}
    for s in scenarios:
        by_spec[s['subcategory']] = by_spec.get(s['subcategory'], 0) + 1
    for k, v in sorted(by_spec.items()):
        print(f'    {k:<8} {v}문항')

    # ── 남은 일 ──
    print('\n' + '=' * 78)
    print('■ 남은 준비')
    print('=' * 78)
    if no_persona:
        print(f'  · 페르소나 나이 미입력 {len(no_persona)}건 — /phr 화면에서 채운다')
        print(f'    (원본에 나이·성별 필드가 없다. 렉스소프트 회신 전이면 임의값을 넣고 그 사실을 명시)')
    print('  · 리허설 — 전 문항을 한 번씩 실행해 인용 수치 오류를 선수정한다')
    print('  · 자문 종료 후 /api/advisory/export 로 판정 전건을 받아 채점 기준을 만든다')
    if args.dry_run:
        print('\n  [DRY RUN] 저장하지 않았습니다. --dry-run 을 빼고 다시 실행하세요.')


if __name__ == '__main__':
    main()
