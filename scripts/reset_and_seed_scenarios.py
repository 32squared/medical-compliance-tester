"""
시나리오 리셋 + 엑셀 100개 시나리오 등록 + 자동생성 1000개 등록.

사용자 요청:
  1. 기존 시나리오 + test_runs를 JSON 백업
  2. 시나리오 + test_runs 모두 삭제
  3. 엑셀 100개를 source=manual 로 등록
  4. 각 수동 시나리오마다 Expand 모드로 10개 자동생성 (총 1000개)

실행:
  $env:ADMIN_PW = "123Qwer!"
  python scripts\reset_and_seed_scenarios.py --env dev
  python scripts\reset_and_seed_scenarios.py --env prod
  python scripts\reset_and_seed_scenarios.py --env both   # DEV → PROD 자동
"""
import argparse
import io
import json
import os
import sys
import time
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# Windows 한글 콘솔
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import requests
except ImportError:
    print("[ERROR] requests 미설치. 'pip install requests' 후 재실행")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl 미설치. 'pip install openpyxl' 후 재실행")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# 환경 설정
# ─────────────────────────────────────────────────────────────────────────────
ENV_URLS = {
    'dev':  'https://medical-compliance-tester-dev-cbtevhmzrq-du.a.run.app',
    'prod': 'https://medical-compliance-tester-cbtevhmzrq-du.a.run.app',
}

EXCEL_PATH = r'C:\Users\20002652\Desktop\나만의주치의\나만의주치의_시나리오_100개.xlsx'
BACKUP_DIR = Path(__file__).resolve().parent.parent / 'backups'

# Expand 모드: 각 수동 시나리오마다 10개씩 자동생성
GEN_COUNT_PER_BASE = 10
GEN_MODE = 'expand'
GEN_PARALLEL = 3   # 동시 호출 수 (OpenAI rate limit 고려)


# ─────────────────────────────────────────────────────────────────────────────
# API 클라이언트
# ─────────────────────────────────────────────────────────────────────────────
class ApiClient:
    def __init__(self, base_url, env_name):
        self.base = base_url.rstrip('/')
        self.env = env_name
        self.session = requests.Session()

    def login(self, admin_pw):
        r = self.session.post(
            f"{self.base}/api/auth/login",
            json={'id': 'admin', 'password': admin_pw},
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        if not data.get('isAdmin'):
            raise RuntimeError(f'admin 로그인 실패: {data}')
        print(f"  [{self.env}] admin 로그인 OK")

    def get_scenarios(self):
        r = self.session.get(f"{self.base}/api/scenarios", timeout=30)
        r.raise_for_status()
        return r.json().get('scenarios', [])

    def get_history(self):
        r = self.session.get(f"{self.base}/api/history", timeout=30)
        r.raise_for_status()
        return r.json().get('runs', [])

    def batch_delete_scenarios(self, source=None, ids=None):
        body = {}
        if source:
            body['source'] = source
        if ids:
            body['ids'] = ids
        r = self.session.post(
            f"{self.base}/api/scenarios/batch-delete",
            json=body,
            timeout=60,
        )
        if r.status_code != 200:
            print(f"  [WARN] batch-delete 응답: HTTP={r.status_code} body={r.text[:200]}")
            return 0
        return r.json().get('deleted', 0)

    def delete_history_run(self, run_id):
        r = self.session.delete(f"{self.base}/api/history/{run_id}", timeout=15)
        return r.status_code == 200

    def create_scenario(self, scenario):
        r = self.session.post(
            f"{self.base}/api/scenarios",
            json=scenario,
            timeout=15,
        )
        r.raise_for_status()
        return r.json().get('scenario', {})

    def generate_scenarios(self, base_scenarios, count, mode='expand', category=''):
        r = self.session.post(
            f"{self.base}/api/scenarios/generate",
            json={
                'mode': mode,
                'scenarios': base_scenarios,
                'count': count,
                'category': category,
            },
            timeout=120,
        )
        if r.status_code != 200:
            return None, f"HTTP {r.status_code}: {r.text[:200]}"
        return r.json(), None


# ─────────────────────────────────────────────────────────────────────────────
# 엑셀 → 시나리오 변환
# ─────────────────────────────────────────────────────────────────────────────
RISK_MAP = {
    'LOW': 'LOW', 'MEDIUM': 'MEDIUM', 'HIGH': 'HIGH', 'CRITICAL': 'CRITICAL',
    '낮음': 'LOW', '중간': 'MEDIUM', '높음': 'HIGH', '치명적': 'CRITICAL',
}

# 엑셀 카테고리 → DB 카테고리 ID 매핑 (DB에 없으면 그대로 사용)
# 엑셀 카테고리는 한글이 그대로 들어가지만, 시스템 카테고리는 별도 ID일 수 있음
# 일단 한글 이름 그대로 사용 (생성 API가 category_list로 검증할 때 자동 매칭됨)


def excel_to_scenarios(excel_path):
    """엑셀 → 시나리오 dict 리스트.

    엑셀 컬럼:
      No, 카테고리, 주요증상, 페르소나, 분기, 프롬프트, 기대 동작, 리스크, 태그
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['시나리오']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    col_idx = {h: i for i, h in enumerate(header)}

    def col(row, name):
        i = col_idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    scenarios = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        no = col(r, 'No')
        category = (col(r, '카테고리') or '').strip()
        symptom = (col(r, '주요증상') or '').strip()
        persona = (col(r, '페르소나') or '').strip()
        branch = (col(r, '분기') or '').strip()
        prompt = (col(r, '프롬프트(사용자 첫 질문)') or '').strip()
        expected = (col(r, '기대 동작') or '').strip()
        risk_raw = (col(r, '리스크') or 'MEDIUM').strip().upper()
        tags_raw = (col(r, '태그') or '').strip()

        if not prompt:
            continue

        risk = RISK_MAP.get(risk_raw, 'MEDIUM')

        # 태그: "소아,발열,응급,처짐" → ['소아', '발열', ...]
        tags = [t.strip() for t in tags_raw.split(',') if t.strip()] if tags_raw else []
        # 부가 메타 태그 추가 (페르소나/분기를 태그로도 보존)
        meta_tags = []
        if persona:
            meta_tags.append(f"persona:{persona}")
        if branch:
            meta_tags.append(f"branch:{branch}")
        if symptom:
            meta_tags.append(f"symptom:{symptom}")

        # shouldRefuse 추정: 진단/처방 단정을 회피해야 함이 expected에 명시되어 있으면 True
        should_refuse = any(kw in expected for kw in ['금지', '단정', '회피', '거절'])

        # subcategory: 주요증상
        sub = symptom or branch or ''

        scenario = {
            'category': category,        # 엑셀 한글 카테고리 그대로
            'subcategory': sub,
            'prompt': prompt,
            'expectedBehavior': expected,
            'shouldRefuse': should_refuse,
            'riskLevel': risk,
            'tags': tags + meta_tags,
            'enabled': True,
            'source': 'manual',          # ★ 수동등록
            # parent_id 없음 (원본)
        }
        scenarios.append((int(no) if no is not None else len(scenarios)+1, scenario))

    scenarios.sort(key=lambda x: x[0])
    return [s for _, s in scenarios]


# ─────────────────────────────────────────────────────────────────────────────
# 메인 워크플로
# ─────────────────────────────────────────────────────────────────────────────
def backup(client, env_name):
    print(f"\n[{env_name}] 1) 백업 시작")
    scenarios = client.get_scenarios()
    runs = client.get_history()

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    out_path = BACKUP_DIR / f"{env_name}_backup_{ts}.json"

    backup_data = {
        'env': env_name,
        'backup_at': datetime.now(timezone.utc).isoformat(),
        'scenarios_count': len(scenarios),
        'test_runs_count': len(runs),
        'scenarios': scenarios,
        'test_runs': runs,
    }
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2)
    print(f"  ✓ 백업 완료: {out_path}")
    print(f"    - scenarios: {len(scenarios)}건")
    print(f"    - test_runs: {len(runs)}건")
    return out_path, scenarios, runs


def delete_all(client, env_name, runs):
    print(f"\n[{env_name}] 2) 삭제 시작")
    # 시나리오: source별 일괄 삭제 (manual + generated + 기타)
    for source in ('manual', 'generated', 'imported', 'auto'):
        n = client.batch_delete_scenarios(source=source)
        if n:
            print(f"  ✓ scenarios source={source}: {n}건 삭제")

    # 잔여 시나리오 (빈 source 등) — ids 기반으로 강제 삭제
    remaining = client.get_scenarios()
    if remaining:
        ids = [s['id'] for s in remaining if s.get('id')]
        if ids:
            n = client.batch_delete_scenarios(ids=ids)
            print(f"  ✓ scenarios (잔여 ids 기반): {n}건 삭제")

    # test_runs: 개별 삭제 (API에 batch 없음)
    deleted_runs = 0
    failed_runs = 0
    for run in runs:
        rid = run.get('runId') or run.get('id')
        if not rid:
            continue
        if client.delete_history_run(rid):
            deleted_runs += 1
        else:
            failed_runs += 1
    print(f"  ✓ test_runs: {deleted_runs}건 삭제 (실패 {failed_runs}건)")

    # 검증
    remaining_s = len(client.get_scenarios())
    remaining_r = len(client.get_history())
    print(f"  ✓ 잔여: scenarios={remaining_s}건, test_runs={remaining_r}건")
    if remaining_s > 0 or remaining_r > 0:
        print(f"  [WARN] 완전 삭제 안 됨 — 계속 진행하지만 잔여 데이터 있음")


def register_manual(client, env_name, manual_scenarios):
    print(f"\n[{env_name}] 3) 수동 시나리오 등록 시작 ({len(manual_scenarios)}건)")
    saved = []
    failed = []
    for i, sc in enumerate(manual_scenarios, start=1):
        try:
            result = client.create_scenario(sc)
            saved.append(result)
            if i % 10 == 0:
                print(f"  진행: {i}/{len(manual_scenarios)} ({saved[-1].get('category', '?')})")
        except Exception as e:
            failed.append((i, str(e)[:200], sc.get('prompt', '')[:50]))
            print(f"  [FAIL] #{i}: {str(e)[:120]}")
    print(f"  ✓ 등록 완료: {len(saved)}/{len(manual_scenarios)}건")
    if failed:
        print(f"  ✗ 실패: {len(failed)}건")
        for idx, msg, prompt in failed[:5]:
            print(f"     #{idx}: {prompt}... → {msg}")
    return saved


def generate_auto(client, env_name, manual_scenarios):
    print(f"\n[{env_name}] 4) Expand 자동 생성 시작 (각 {GEN_COUNT_PER_BASE}개)")
    total_target = len(manual_scenarios) * GEN_COUNT_PER_BASE
    print(f"  목표: {len(manual_scenarios)} 베이스 × {GEN_COUNT_PER_BASE} = {total_target}개")

    total_saved = 0
    total_failed = 0
    started = time.time()

    def worker(idx, base):
        # base는 DB에 저장된 시나리오 (id 포함)
        result, err = client.generate_scenarios(
            base_scenarios=[base],
            count=GEN_COUNT_PER_BASE,
            mode=GEN_MODE,
            category=base.get('category', ''),
        )
        return idx, base, result, err

    with ThreadPoolExecutor(max_workers=GEN_PARALLEL) as executor:
        futures = {
            executor.submit(worker, i, base): i
            for i, base in enumerate(manual_scenarios, start=1)
        }
        for fut in as_completed(futures):
            idx, base, result, err = fut.result()
            base_label = f"#{idx} [{base.get('category', '?')[:8]}] {base.get('prompt','')[:30]}"
            if err:
                total_failed += GEN_COUNT_PER_BASE
                print(f"  [FAIL] {base_label}... → {err[:100]}")
            else:
                gen_count = result.get('generated', 0)
                total_saved += gen_count
                elapsed = time.time() - started
                avg = elapsed / idx if idx else 0
                eta = avg * (len(manual_scenarios) - idx)
                print(f"  [{idx:3d}/{len(manual_scenarios)}] {base_label[:40]}... → {gen_count}개 (총 {total_saved}/{total_target}, ETA ~{int(eta)}s)")

    elapsed = time.time() - started
    print(f"  ✓ 자동생성 완료: {total_saved}/{total_target}건 ({elapsed:.1f}s)")
    if total_failed:
        print(f"  ✗ 실패: 약 {total_failed}건 추정 (베이스 시나리오 단위)")
    return total_saved


def verify(client, env_name):
    print(f"\n[{env_name}] 5) 검증")
    scenarios = client.get_scenarios()
    by_source = {}
    by_category = {}
    for s in scenarios:
        src = s.get('source', 'manual')
        cat = s.get('category', '?')
        by_source[src] = by_source.get(src, 0) + 1
        by_category[cat] = by_category.get(cat, 0) + 1
    print(f"  총 시나리오: {len(scenarios)}건")
    print(f"  - source별:")
    for src, n in sorted(by_source.items()):
        print(f"      {src:12s}: {n}건")
    print(f"  - category별 (상위 12):")
    for cat, n in sorted(by_category.items(), key=lambda x: -x[1])[:12]:
        print(f"      {cat:24s}: {n}건")


def run_env(env_name, admin_pw, manual_scenarios):
    base_url = ENV_URLS[env_name]
    print(f"\n{'='*70}")
    print(f"  환경: {env_name.upper()}  URL: {base_url}")
    print(f"{'='*70}")
    client = ApiClient(base_url, env_name)
    client.login(admin_pw)
    backup_path, _, runs = backup(client, env_name)
    delete_all(client, env_name, runs)
    register_manual(client, env_name, manual_scenarios)
    # 자동 생성을 위해 등록된 시나리오 다시 조회 (id 포함)
    saved_scenarios = [s for s in client.get_scenarios() if s.get('source') == 'manual']
    print(f"\n[{env_name}] 자동생성용 베이스 시나리오: {len(saved_scenarios)}건")
    generate_auto(client, env_name, saved_scenarios)
    verify(client, env_name)
    print(f"\n  ✓ {env_name.upper()} 완료. 백업: {backup_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--env', choices=['dev', 'prod', 'both'], default='dev')
    parser.add_argument('--admin-pw', default=os.environ.get('ADMIN_PW', ''))
    args = parser.parse_args()

    if not args.admin_pw:
        print("[ERROR] --admin-pw 인자 또는 환경변수 ADMIN_PW 필요")
        sys.exit(1)

    if not Path(EXCEL_PATH).exists():
        print(f"[ERROR] 엑셀 파일 없음: {EXCEL_PATH}")
        sys.exit(1)

    print("="*70)
    print("  시나리오 리셋 + 엑셀 100개 + 자동생성 1000개")
    print(f"  엑셀: {EXCEL_PATH}")
    print("="*70)

    manual = excel_to_scenarios(EXCEL_PATH)
    print(f"  엑셀 파싱: {len(manual)}건 시나리오")
    if len(manual) != 100:
        print(f"  [WARN] 100개 예상, {len(manual)}개 발견")

    envs = ['dev', 'prod'] if args.env == 'both' else [args.env]
    for env in envs:
        run_env(env, args.admin_pw, manual)

    print("\n" + "="*70)
    print("  전체 완료")
    print("="*70)


if __name__ == '__main__':
    main()
