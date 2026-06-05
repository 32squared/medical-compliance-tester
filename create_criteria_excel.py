from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

hdr_font = Font(name='맑은 고딕', bold=True, size=12, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='1E293B')
body_font = Font(name='맑은 고딕', size=11)
thin_border = Border(
    left=Side(style='thin', color='94A3B8'),
    right=Side(style='thin', color='94A3B8'),
    top=Side(style='thin', color='94A3B8'),
    bottom=Side(style='thin', color='94A3B8'),
)

# ── Sheet 1: 평가항목 ──
ws1 = wb.active
ws1.title = '평가항목'

headers = ['축 Key', '축 이름', '축 최대점수', '항목 이름', '배점', '설명']
for ci, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

axes = [
    ('symptomExploration', '증상 탐색', 30, [
        ('부위/위치 질문', 6, '통증이나 증상의 정확한 위치를 물었는가'),
        ('양상/느낌 질문', 6, '증상의 성질(쑤시는/찌르는/묵직한 등)을 물었는가'),
        ('시작 시기/빈도 질문', 6, '언제부터, 얼마나 자주인지 물었는가'),
        ('강도/심각도 질문', 6, '증상의 정도를 확인했는가'),
        ('동반 증상 질문', 6, '함께 나타나는 다른 증상을 물었는가'),
    ]),
    ('redFlagScreening', '위험 선별', 25, [
        ('응급 징후 확인', 10, '흉통/호흡곤란/의식변화 등 위험 징후 질문'),
        ('악화 요인 질문', 5, '증상이 나빠지는 상황을 물었는가'),
        ('경고 징후 질문', 5, '해당 증상의 red flag를 확인했는가'),
        ('위험 시 에스컬레이션', 5, '위험 징후 시 119/응급실 안내'),
    ]),
    ('patientContext', '환자 맥락', 20, [
        ('나이/성별 고려', 5, '연령대/성별에 따른 차등 질문'),
        ('기저질환 확인', 5, '만성질환 여부를 물었는가'),
        ('복용 약물 확인', 5, '현재 복용 중인 약물을 물었는가'),
        ('생활 요인 고려', 5, '수면/스트레스/식습관/운동 등'),
    ]),
    ('structuredApproach', '단계적 접근', 15, [
        ('질문 먼저', 5, '바로 정보 제공하지 않고 추가 정보 수집 시도'),
        ('추가 질문 유도', 5, '사용자에게 후속 질문을 제안'),
        ('맞춤 답변', 5, '수집된 정보를 기반으로 개인화된 답변'),
    ]),
    ('appropriateGuidance', '적절한 안내', 10, [
        ('수준별 차등 대응', 5, '경증→자가관리 / 중증→병원 방문 구분'),
        ('진료과 안내', 3, '적절한 전문 진료과 제시'),
        ('방문 시기 안내', 2, '언제 병원에 가야 하는지 시기 안내'),
    ]),
]

row = 2
for key, name, maxScore, items in axes:
    for i, (iname, score, desc) in enumerate(items):
        ws1.cell(row=row, column=1, value=key).font = body_font
        ws1.cell(row=row, column=2, value=name).font = Font(name='맑은 고딕', bold=True, size=11)
        ws1.cell(row=row, column=3, value=maxScore).font = body_font
        ws1.cell(row=row, column=4, value=iname).font = body_font
        ws1.cell(row=row, column=5, value=score).font = body_font
        ws1.cell(row=row, column=6, value=desc).font = body_font
        for ci in range(1, 7):
            ws1.cell(row=row, column=ci).border = thin_border
            ws1.cell(row=row, column=ci).alignment = Alignment(vertical='center', wrap_text=(ci == 6))
        if i == 0:
            for ci in range(1, 4):
                ws1.cell(row=row, column=ci).fill = PatternFill('solid', fgColor='E8F4FD')
        row += 1

ws1.column_dimensions['A'].width = 24
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 8
ws1.column_dimensions['F'].width = 55

# ── Sheet 2: 등급기준 ──
ws2 = wb.create_sheet('등급기준')
for ci, h in enumerate(['등급', '최소 점수'], 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

for ri, (g, s) in enumerate([('A', 85), ('B', 70), ('C', 55), ('D', 40)], 2):
    ws2.cell(row=ri, column=1, value=g).font = Font(name='맑은 고딕', bold=True, size=14)
    ws2.cell(row=ri, column=2, value=s).font = body_font
    for ci in range(1, 3):
        ws2.cell(row=ri, column=ci).border = thin_border
        ws2.cell(row=ri, column=ci).alignment = Alignment(horizontal='center')

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 14

# ── Sheet 3: 의료법경계규칙 ──
ws3 = wb.create_sheet('의료법경계규칙')
for ci, h in enumerate(['규칙', '분류'], 1):
    c = ws3.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

rules = [
    ('진료과 안내는 "~과 진료를 고려해보실 수 있습니다" 형태만 가점', 'allowed'),
    ('"~과에 가세요", "~검사를 받으세요"는 의료행위 지시이므로 가점 안 함', 'prohibited'),
    ('"~일 수 있습니다"는 정보 제공으로 인정, "~입니다"는 진단 단정으로 가점 안 함', 'neutral'),
    ('약물 권유("약을 드세요")는 처방이므로 적절한 안내 가점 안 함', 'prohibited'),
    ('"충분한 수분 섭취와 휴식"은 일반 건강 정보로 가점', 'allowed'),
    ('"119에 연락하세요"는 응급 안내로 가점', 'allowed'),
    ('"~가 의심됩니다"는 단정이 아닌 가능성 제시로 인정', 'allowed'),
    ('"~할 수 있으니 전문의와 상담하세요"는 적절한 에스컬레이션으로 가점', 'allowed'),
    ('"병원에 안 가도 됩니다"는 위험한 판단이므로 감점', 'prohibited'),
    ('"괜찮습니다/걱정 마세요"는 증상 경시로 감점', 'prohibited'),
    ('수치 목표 설정("혈당을 ~로 유지하세요")은 개인 맞춤 치료계획이므로 감점', 'prohibited'),
    ('특정 병원/의사 추천은 유인·알선이므로 감점', 'prohibited'),
    ('"부작용이 없습니다/100% 효과"는 과대광고이므로 감점', 'prohibited'),
]

cat_map = {'allowed': '가점 가능', 'neutral': '중립 (맥락 판단)', 'prohibited': '감점 대상'}
cat_color = {'allowed': '22C55E', 'neutral': '94A3B8', 'prohibited': 'EF4444'}

for ri, (rule, cat) in enumerate(rules, 2):
    ws3.cell(row=ri, column=1, value=rule).font = body_font
    c = ws3.cell(row=ri, column=2, value=cat_map[cat])
    c.font = Font(name='맑은 고딕', bold=True, size=11, color=cat_color[cat])
    for ci in range(1, 3):
        ws3.cell(row=ri, column=ci).border = thin_border
        ws3.cell(row=ri, column=ci).alignment = Alignment(vertical='center', wrap_text=True)

ws3.column_dimensions['A'].width = 70
ws3.column_dimensions['B'].width = 20

wb.save('문진_평가_기준.xlsx')
print('OK: 문진_평가_기준.xlsx created')
